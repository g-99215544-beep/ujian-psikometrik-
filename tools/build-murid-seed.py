from pathlib import Path
import json
import re

import openpyxl


SOURCE = Path("PBB1013 Keseluruhan Murid as of 2026-04-28.xlsx")
OUTPUT = Path("firebase-seed/ujianpsikometrikapp.seed.json")
ROOT = "ujianpsikometrikapp"
TARGET_YEARS = {
    "TAHUN EMPAT": 4,
    "TAHUN LIMA": 5,
    "TAHUN ENAM": 6,
}


def clean_key(value):
    return re.sub(r"[^0-9A-Za-z]", "", str(value or "")).upper()


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    murid_by_ic = {}
    skipped = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tahun_label = str(row[col["TAHUN / TINGKATAN"]] or "").strip().upper()
        if tahun_label not in TARGET_YEARS:
            continue

        key = clean_key(row[col["NO. PENGENALAN"]])
        nama = str(row[col["NAMA"]] or "").strip()
        nama_kelas = str(row[col["NAMA KELAS"]] or "").strip().upper()
        tahun = TARGET_YEARS[tahun_label]

        if not key or not nama or not nama_kelas:
            skipped.append({"row": row_num, "reason": "missing required field"})
            continue

        murid_by_ic[key] = {
            "nama": nama,
            "kelas": f"{tahun} {nama_kelas}",
            "tahun": tahun,
            "namaKelas": nama_kelas,
            "noPengenalan": key,
            "sekolah": "SK Sri Aman",
        }

    payload = {
        ROOT: {
            "meta": {
                "sumber": SOURCE.name,
                "tahap": "Tahap 2",
                "tahun": [4, 5, 6],
                "jumlahMurid": len(murid_by_ic),
                "skipped": skipped,
            },
            "muridByIc": murid_by_ic,
        }
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(murid_by_ic)} students to {OUTPUT}")
    if skipped:
        print(f"Skipped {len(skipped)} rows")


if __name__ == "__main__":
    main()
